"""
Image Conversion Tools — Image to PDF and Image to Excel (OCR).

Combined into a single module since they share OCR dependencies and
both operate on image inputs.

- Image to PDF: Uses Pillow for processing and fpdf2 for generation.
- Image to Excel: Uses Pillow for preprocessing, pytesseract for OCR,
  and openpyxl for Excel output.
"""

import logging
from pathlib import Path
from PIL import Image, ImageOps
from fpdf import FPDF
import pytesseract
from pytesseract import Output
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from backend.config import TESSERACT_PATH
from backend.converters import ConversionError

logger = logging.getLogger(__name__)

try:
    import openpyxl.utils
except ImportError:
    openpyxl = None

# ─── Image to PDF Constants ─────────────────────────────────────────────────
_A4_WIDTH_MM = 210.0
_A4_HEIGHT_MM = 297.0
_PAGE_MARGIN_MM = 10.0
_MAX_IMAGE_PIXELS = 4000

# ─── Image to Excel Constants ───────────────────────────────────────────────
_TESSERACT_CONFIG = "--oem 3 --psm 3"
_MIN_CONFIDENCE = 30
_MIN_IMAGE_DIMENSION = 1000
_MAX_IMAGE_DIMENSION = 4000
_ROW_Y_THRESHOLD_FACTOR = 0.5
_COLUMN_GAP_MULTIPLIER = 3.0
_MIN_COLUMN_GAP_PX = 20
_COLUMN_GAP_WORD_WIDTH_FACTOR = 0.8
_HEADER_FILL_COLOR = "4472C4"
_HEADER_FONT_COLOR = "FFFFFF"
_MAX_COLUMN_WIDTH_CHARS = 50


# ═══════════════════════════════════════════════════════════════════════════════
# Image to PDF
# ═══════════════════════════════════════════════════════════════════════════════

def convert_images_to_pdf(
    image_paths: list[Path],
    output_path: Path,
) -> tuple[Path, int]:
    """Convert a list of images into a single multi-page PDF."""
    if not image_paths:
        raise ConversionError("No images provided for conversion.")

    for i, path in enumerate(image_paths):
        if not path.exists():
            raise ConversionError(f"Image file #{i + 1} was not found.")
        if path.stat().st_size == 0:
            raise ConversionError(f"Image file #{i + 1} ('{path.name}') is empty.")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    logger.info("Image→PDF: %d images", len(image_paths))

    try:
        pdf = FPDF(unit="mm", format="A4")
        pdf.set_auto_page_break(auto=False)
        pdf.set_compression_level(6)

        for idx, img_path in enumerate(image_paths):
            _add_image_page(pdf, img_path, idx + 1, len(image_paths))

        pdf.output(str(output_path))
    except ConversionError:
        raise
    except Exception as e:
        logger.error("Image→PDF generation error: %s", e, exc_info=True)
        raise ConversionError("Failed to generate the PDF.", detail=str(e))

    if not output_path.exists() or output_path.stat().st_size == 0:
        raise ConversionError("PDF generation failed — output is empty.")

    logger.info("Image→PDF complete: %d pages → %s (%.1f KB)", len(image_paths), output_path.name, output_path.stat().st_size / 1024)
    return output_path, len(image_paths)


def _add_image_page(pdf: FPDF, img_path: Path, index: int, total: int) -> None:
    """Open, preprocess, and embed a single image as a PDF page."""
    try:
        img = Image.open(img_path)
    except Exception as e:
        raise ConversionError(f"Cannot open image #{index} ('{img_path.name}').", detail=str(e))

    try:
        if img.mode in ("RGBA", "LA", "PA"):
            if img.mode == "PA":
                img = img.convert("RGBA")
            background = Image.new("RGB", img.size, (255, 255, 255))
            background.paste(img, mask=img.split()[-1])
            img = background
        elif img.mode == "P":
            img_rgba = img.convert("RGBA")
            background = Image.new("RGB", img.size, (255, 255, 255))
            background.paste(img_rgba, mask=img_rgba.split()[-1])
            img = background
        elif img.mode != "RGB":
            img = img.convert("RGB")

        img_w_px, img_h_px = img.size
        max_dim = max(img_w_px, img_h_px)
        if max_dim > _MAX_IMAGE_PIXELS:
            scale = _MAX_IMAGE_PIXELS / max_dim
            img = img.resize((int(img_w_px * scale), int(img_h_px * scale)), Image.LANCZOS)
            img_w_px, img_h_px = img.size

        aspect = img_w_px / img_h_px
        if aspect > (_A4_WIDTH_MM / _A4_HEIGHT_MM):
            orientation, page_w, page_h = "L", _A4_HEIGHT_MM, _A4_WIDTH_MM
        else:
            orientation, page_w, page_h = "P", _A4_WIDTH_MM, _A4_HEIGHT_MM

        avail_w, avail_h = page_w - 2 * _PAGE_MARGIN_MM, page_h - 2 * _PAGE_MARGIN_MM
        if aspect > (avail_w / avail_h):
            display_w, display_h = avail_w, avail_w / aspect
        else:
            display_h, display_w = avail_h, avail_h * aspect

        x = _PAGE_MARGIN_MM + (avail_w - display_w) / 2
        y = _PAGE_MARGIN_MM + (avail_h - display_h) / 2

        pdf.add_page(orientation=orientation)
        pdf.image(img, x=x, y=y, w=display_w, h=display_h)
    finally:
        img.close()


# ═══════════════════════════════════════════════════════════════════════════════
# Image to Excel (OCR)
# ═══════════════════════════════════════════════════════════════════════════════

def convert_image_to_excel(
    input_path: Path,
    output_path: Path,
) -> tuple[Path, int]:
    """Extract tabular data from an image using OCR and write to Excel."""
    _verify_tesseract()

    if not input_path.exists():
        raise ConversionError("The uploaded image file was not found.")
    if input_path.stat().st_size == 0:
        raise ConversionError("The uploaded image file is empty (0 bytes).")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    logger.info("Image→Excel: %s (%.1f KB)", input_path.name, input_path.stat().st_size / 1024)

    try:
        img = Image.open(input_path)
        processed_img = _preprocess_image(img)
        words = _extract_words(processed_img)

        if not words:
            raise ConversionError(
                "No text was detected in this image. Ensure it contains clear, printed text."
            )

        logger.info("OCR detected %d word(s)", len(words))
        table = _build_table(words)

        if not table or (len(table) == 1 and len(table[0]) == 1):
            raise ConversionError(
                "Could not detect a table structure. Try uploading a clearer image of a table."
            )

        logger.info("Table built: %d rows x %d columns", len(table), max(len(row) for row in table))
        data_rows = _write_excel(table, output_path)

    except ConversionError:
        raise
    except Exception as e:
        logger.error("Image→Excel error: %s", e, exc_info=True)
        raise ConversionError("Failed to extract data from the image.", detail=str(e))
    finally:
        img.close()

    if not output_path.exists() or output_path.stat().st_size == 0:
        raise ConversionError("Excel generation failed — output is empty.")

    logger.info("Image→Excel complete: %d rows → %s (%.1f KB)", data_rows, output_path.name, output_path.stat().st_size / 1024)
    return output_path, data_rows


def _preprocess_image(img: Image.Image) -> Image.Image:
    """Grayscale, resize, and boost contrast for optimal OCR."""
    if img.mode != "RGB":
        img = img.convert("RGB")
    gray = img.convert("L")
    w, h = gray.size

    min_dim, max_dim = min(w, h), max(w, h)
    if min_dim < _MIN_IMAGE_DIMENSION:
        scale = _MIN_IMAGE_DIMENSION / min_dim
        gray = gray.resize((int(w * scale), int(h * scale)), Image.LANCZOS)
        w, h = gray.size
    max_dim = max(w, h)
    if max_dim > _MAX_IMAGE_DIMENSION:
        scale = _MAX_IMAGE_DIMENSION / max_dim
        gray = gray.resize((int(w * scale), int(h * scale)), Image.LANCZOS)
        
    return ImageOps.autocontrast(gray, cutoff=2)


def _extract_words(img: Image.Image) -> list[dict]:
    """Run Tesseract word-level detection and filter by confidence."""
    try:
        data = pytesseract.image_to_data(img, output_type=Output.DICT, config=_TESSERACT_CONFIG)
    except pytesseract.TesseractNotFoundError:
        raise ConversionError("OCR is unavailable. Tesseract is not installed.", detail=f"Path: {TESSERACT_PATH}")
    except Exception as e:
        raise ConversionError("Tesseract failed to process the image.", detail=str(e))

    words = []
    for i in range(len(data["text"])):
        text = data["text"][i].strip()
        try:
            conf = int(data["conf"][i])
        except (ValueError, TypeError):
            continue
        if not text or conf < _MIN_CONFIDENCE:
            continue
        words.append({
            "text": text, "left": int(data["left"][i]), "top": int(data["top"][i]),
            "width": int(data["width"][i]), "height": int(data["height"][i]), "conf": conf,
        })
    return words


def _build_table(words: list[dict]) -> list[list[str]]:
    """Convert flat word list into 2D table via row grouping and column splitting."""
    rows = _group_into_rows(words)
    if not rows:
        return []

    table = [_split_row_into_cells(row_words) for row_words in rows]
    max_cols = max(len(row) for row in table)
    for row in table:
        while len(row) < max_cols:
            row.append("")
    return table


def _group_into_rows(words: list[dict]) -> list[list[dict]]:
    """Group words into rows by Y-coordinate proximity."""
    if not words:
        return []
    heights = sorted(w["height"] for w in words)
    median_height = heights[len(heights) // 2]
    y_threshold = median_height * _ROW_Y_THRESHOLD_FACTOR
    sorted_words = sorted(words, key=lambda w: (w["top"], w["left"]))
    
    rows, current_row = [], [sorted_words[0]]
    for word in sorted_words[1:]:
        current_tops = sorted(w["top"] for w in current_row)
        row_median_y = current_tops[len(current_tops) // 2]
        if abs(word["top"] - row_median_y) <= y_threshold:
            current_row.append(word)
        else:
            current_row.sort(key=lambda w: w["left"])
            rows.append(current_row)
            current_row = [word]
    if current_row:
        current_row.sort(key=lambda w: w["left"])
        rows.append(current_row)
    return rows


def _split_row_into_cells(row_words: list[dict]) -> list[str]:
    """Split a row of words into cells by detecting horizontal gaps."""
    if not row_words:
        return []
    if len(row_words) == 1:
        return [row_words[0]["text"]]

    gaps = []
    for i in range(len(row_words) - 1):
        right_edge = row_words[i]["left"] + row_words[i]["width"]
        gaps.append(max(0.0, float(row_words[i + 1]["left"] - right_edge)))

    if not gaps:
        return [" ".join(w["text"] for w in row_words)]

    sorted_gaps = sorted(gaps)
    median_gap = sorted_gaps[len(sorted_gaps) // 2]
    avg_word_width = sum(w["width"] for w in row_words) / len(row_words)
    threshold = max(median_gap * _COLUMN_GAP_MULTIPLIER, avg_word_width * _COLUMN_GAP_WORD_WIDTH_FACTOR, _MIN_COLUMN_GAP_PX)

    cells, current_cell_words = [], [row_words[0]["text"]]
    for i, gap in enumerate(gaps):
        if gap > threshold:
            cells.append(" ".join(current_cell_words))
            current_cell_words = [row_words[i + 1]["text"]]
        else:
            current_cell_words.append(row_words[i + 1]["text"])
    if current_cell_words:
        cells.append(" ".join(current_cell_words))
    return cells


def _write_excel(table: list[list[str]], output_path: Path) -> int:
    """Write 2D table to formatted Excel file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Extracted Data"

    for row_idx, row_data in enumerate(table, start=1):
        for col_idx, cell_value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    has_header = _looks_like_header(table)
    if has_header and table:
        header_fill = PatternFill(start_color=_HEADER_FILL_COLOR, end_color=_HEADER_FILL_COLOR, fill_type="solid")
        header_font = Font(bold=True, color=_HEADER_FONT_COLOR, size=11)
        for col_idx in range(1, len(table[0]) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill, cell.font = header_fill, header_font
            cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"), right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"), bottom=Side(style="thin", color="D9D9D9"),
    )
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border

    for col_idx in range(1, ws.max_column + 1):
        max_len = max((len(str(c.value)) for c in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=1) if c.value), default=0)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(min(max_len + 3, _MAX_COLUMN_WIDTH_CHARS), 8)

    wb.save(str(output_path))
    return max(0, len(table) - 1 if has_header else len(table))


def _looks_like_header(table: list[list[str]]) -> bool:
    if len(table) < 2 or len(table[0]) < 2:
        return False
    if not all(c.strip() for c in table[0]):
        return False
    first_avg = sum(len(c) for c in table[0]) / len(table[0])
    remaining = [len(c) for row in table[1:] for c in row if c.strip()]
    return remaining and first_avg < (sum(remaining) / len(remaining))


def _verify_tesseract() -> None:
    if TESSERACT_PATH != "tesseract":
        pytesseract.pytesseract.tesseract_cmd = TESSERACT_PATH
