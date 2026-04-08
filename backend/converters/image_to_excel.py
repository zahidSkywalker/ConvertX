"""
Image to Excel converter — extracts tabular data from images using OCR.

Pipeline:
  1. Preprocess the image (grayscale, resize, contrast boost) for better OCR.
  2. Run Tesseract OCR in word-level detection mode to get bounding boxes.
  3. Group detected words into rows by Y-coordinate clustering.
  4. Detect column boundaries by analyzing horizontal gaps between words.
  5. Assemble a 2D table from the row/column assignments.
  6. Write the table to an .xlsx file with auto-sized columns.

This works best with:
  - Clear, high-contrast table images (screenshots, scans of printed tables).
  - Straight (non-rotated) tables with visible grid lines or clear column gaps.
  - Images at 200+ DPI (lower resolution images are upscaled during preprocessing).

Limitations:
  - Handwritten text recognition is poor (Tesseract is optimized for print).
  - Merged cells, nested tables, and rotated text are not handled.
  - Tables without clear column separators may produce incorrect column splits.
  - Non-Latin scripts require additional Tesseract language packs.
"""

import logging
from pathlib import Path

from PIL import Image, ImageOps
import pytesseract
from pytesseract import Output
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

from backend.config import TESSERACT_PATH
from backend.converters import ConversionError

logger = logging.getLogger(__name__)

# ─── Tesseract configuration ────────────────────────────────────────────────
# OEM 3 = LSTM neural network (best accuracy, default)
# PSM 3 = Fully automatic page segmentation (good word-level detection)
_TESSERACT_CONFIG = "--oem 3 --psm 3"

# Minimum OCR confidence (0-100) to include a word in the output.
# Below ~30, Tesseract is essentially guessing and produces garbage.
_MIN_CONFIDENCE = 30

# Image preprocessing parameters
_MIN_IMAGE_DIMENSION = 1000   # Resize if shortest side is below this (pixels)
_MAX_IMAGE_DIMENSION = 4000   # Downscale if longest side exceeds this (pixels)

# Row grouping: words within this fraction of the median word height
# are considered to be on the same row.
_ROW_Y_THRESHOLD_FACTOR = 0.5

# Column gap detection: a gap larger than this multiple of the median
# intra-row gap is treated as a column separator.
_COLUMN_GAP_MULTIPLIER = 3.0

# Minimum column gap in pixels (absolute floor, prevents false splits
# on tightly-spaced text).
_MIN_COLUMN_GAP_PX = 20

# Column gap also scaled by average word width — a gap wider than
# this fraction of the average word width is a column boundary.
_COLUMN_GAP_WORD_WIDTH_FACTOR = 0.8

# Excel formatting
_HEADER_FILL_COLOR = "4472C4"  # Blue
_HEADER_FONT_COLOR = "FFFFFF"  # White
_MAX_COLUMN_WIDTH_CHARS = 50


def convert_image_to_excel(
    input_path: Path,
    output_path: Path,
) -> tuple[Path, int]:
    """
    Extract tabular data from an image using OCR and write it to an Excel file.

    Args:
        input_path: Path to the source image file.
        output_path: Path where the output .xlsx file will be written.
                     Parent directory must exist.

    Returns:
        Tuple of (output_file_path, number_of_data_rows_extracted).
        Data rows excludes the header row (if detected).

    Raises:
        ConversionError: If Tesseract is not installed, the image cannot be
                         opened, no text is detected, or Excel generation fails.
    """
    # ── Check Tesseract availability ──
    _verify_tesseract()

    # ── Validate inputs ──
    if not input_path.exists():
        raise ConversionError(
            message="The uploaded image file was not found.",
            detail=f"Expected file at: {input_path}",
        )

    if input_path.stat().st_size == 0:
        raise ConversionError(
            message="The uploaded image file is empty (0 bytes).",
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)

    logger.info(
        "Starting Image→Excel: %s (%.1f KB)",
        input_path.name,
        input_path.stat().st_size / 1024,
    )

    # ── Open image ──
    try:
        img = Image.open(input_path)
    except Exception as e:
        raise ConversionError(
            message="Cannot open the uploaded image. The file may be corrupted.",
            detail=str(e),
        )

    try:
        # ── Preprocess for OCR ──
        processed_img = _preprocess_image(img)

        # ── Run OCR — word-level detection ──
        words = _extract_words(processed_img)

        if not words:
            raise ConversionError(
                message="No text was detected in this image. "
                        "Please ensure the image contains clear, printed text "
                        "with sufficient contrast.",
            )

        logger.info("OCR detected %d word(s) in the image", len(words))

        # ── Build table from word positions ──
        table = _build_table(words)

        if not table or (len(table) == 1 and len(table[0]) == 1):
            # Either no rows, or a single cell — not useful as a table
            raise ConversionError(
                message="Could not detect a table structure in this image. "
                        "The text was detected but could not be organized into "
                        "rows and columns. Try uploading a clearer image of a table.",
            )

        logger.info(
            "Table built: %d rows x %d columns",
            len(table),
            max(len(row) for row in table),
        )

        # ── Write Excel file ──
        data_rows = _write_excel(table, output_path)

    except ConversionError:
        raise
    except Exception as e:
        logger.error("Error in Image→Excel pipeline: %s", e, exc_info=True)
        raise ConversionError(
            message="Failed to extract data from the image.",
            detail=f"{type(e).__name__}: {e}",
        )
    finally:
        img.close()

    # ── Verify output ──
    if not output_path.exists() or output_path.stat().st_size == 0:
        raise ConversionError(
            message="Excel file generation failed — output is empty or missing.",
        )

    logger.info(
        "Image→Excel complete: %d rows → %s (%.1f KB)",
        data_rows,
        output_path.name,
        output_path.stat().st_size / 1024,
    )

    return output_path, data_rows


# ═══════════════════════════════════════════════════════════════════════════════
# Preprocessing
# ═══════════════════════════════════════════════════════════════════════════════

def _preprocess_image(img: Image.Image) -> Image.Image:
    """
    Prepare an image for OCR by converting to grayscale, resizing,
    and boosting contrast.

    The goal is to produce an image where text is clearly distinguishable
    from the background, without introducing artifacts that confuse Tesseract.
    """
    # Convert to RGB first (handles all modes including RGBA, P, LA)
    if img.mode != "RGB":
        img = img.convert("RGB")

    # Convert to grayscale
    gray = img.convert("L")

    w, h = gray.size

    # ── Resize if too small ──
    min_dim = min(w, h)
    if min_dim < _MIN_IMAGE_DIMENSION:
        scale = _MIN_IMAGE_DIMENSION / min_dim
        new_w = int(w * scale)
        new_h = int(h * scale)
        gray = gray.resize((new_w, new_h), Image.LANCZOS)
        w, h = new_w, new_h
        logger.debug("Resized up: %dx%d → %dx%d", w, h, new_w, new_h)

    # ── Downscale if too large (performance) ──
    max_dim = max(w, h)
    if max_dim > _MAX_IMAGE_DIMENSION:
        scale = _MAX_IMAGE_DIMENSION / max_dim
        new_w = int(w * scale)
        new_h = int(h * scale)
        gray = gray.resize((new_w, new_h), Image.LANCZOS)
        w, h = new_w, new_h
        logger.debug("Resized down: %dx%d → %dx%d", w, h, new_w, new_h)

    # ── Boost contrast ──
    # Autocontrast stretches the histogram to use the full 0-255 range.
    # The cutoff parameter trims the darkest/lightest 2% of pixels,
    # which removes noise from scanner artifacts and JPEG compression.
    gray = ImageOps.autocontrast(gray, cutoff=2)

    return gray


# ═══════════════════════════════════════════════════════════════════════════════
# OCR Word Extraction
# ═══════════════════════════════════════════════════════════════════════════════

def _extract_words(img: Image.Image) -> list[dict]:
    """
    Run Tesseract OCR on a preprocessed image and extract word-level data.

    Returns a list of dicts, each with:
        text: str     — the recognized word
        left: int     — x coordinate of the word's left edge (pixels)
        top: int      — y coordinate of the word's top edge (pixels)
        width: int    — width of the word bounding box (pixels)
        height: int   — height of the word bounding box (pixels)
        conf: int     — confidence score (0-100)

    Words below the minimum confidence threshold are filtered out.
    """
    try:
        data = pytesseract.image_to_data(
            img,
            output_type=Output.DICT,
            config=_TESSERACT_CONFIG,
        )
    except pytesseract.TesseractNotFoundError:
        raise ConversionError(
            message="OCR is not available on this server. "
                    "Tesseract is not installed.",
            detail=f"Tesseract command: {TESSERACT_PATH}",
        )
    except Exception as e:
        raise ConversionError(
            message="Tesseract OCR failed to process the image.",
            detail=f"{type(e).__name__}: {e}",
        )

    words = []
    for i in range(len(data["text"])):
        text = data["text"][i].strip()
        try:
            conf = int(data["conf"][i])
        except (ValueError, TypeError):
            continue

        # Skip empty text blocks and low-confidence guesses
        if not text or conf < _MIN_CONFIDENCE:
            continue

        words.append({
            "text": text,
            "left": int(data["left"][i]),
            "top": int(data["top"][i]),
            "width": int(data["width"][i]),
            "height": int(data["height"][i]),
            "conf": conf,
        })

    return words


# ═══════════════════════════════════════════════════════════════════════════════
# Table Construction
# ═══════════════════════════════════════════════════════════════════════════════

def _build_table(words: list[dict]) -> list[list[str]]:
    """
    Convert a flat list of word detections into a 2D table.

    Algorithm:
      1. Group words into rows by Y-coordinate proximity.
      2. For each row, detect column boundaries from horizontal gaps.
      3. Split each row into cells at column boundaries.
      4. Normalize column count across rows (pad with empty strings).
    """
    # ── Step 1: Group into rows ──
    rows = _group_into_rows(words)
    if not rows:
        return []

    logger.debug("Grouped %d words into %d rows", len(words), len(rows))

    # ── Step 2 & 3: Split each row into cells ──
    table = []
    for row_words in rows:
        cells = _split_row_into_cells(row_words)
        table.append(cells)

    # ── Step 4: Normalize column count ──
    max_cols = max(len(row) for row in table)
    if max_cols <= 1:
        # Single-column result — not really a table, but return it anyway
        # The caller decides whether to reject it
        pass

    for row in table:
        while len(row) < max_cols:
            row.append("")

    return table


def _group_into_rows(words: list[dict]) -> list[list[dict]]:
    """
    Group words into rows based on Y-coordinate proximity.

    Uses an adaptive threshold derived from the median word height:
    words whose top coordinate is within half a line-height of the
    current row's median Y are assigned to the same row.
    """
    if not words:
        return []

    # Calculate adaptive threshold from median word height
    heights = sorted(w["height"] for w in words)
    median_height = heights[len(heights) // 2]
    y_threshold = median_height * _ROW_Y_THRESHOLD_FACTOR

    # Sort by Y first, then X (for consistent processing)
    sorted_words = sorted(words, key=lambda w: (w["top"], w["left"]))

    rows: list[list[dict]] = []
    current_row = [sorted_words[0]]

    for word in sorted_words[1:]:
        # Compute the median Y of the current row for comparison
        current_tops = sorted(w["top"] for w in current_row)
        row_median_y = current_tops[len(current_tops) // 2]

        if abs(word["top"] - row_median_y) <= y_threshold:
            current_row.append(word)
        else:
            # Finalize current row: sort by X
            current_row.sort(key=lambda w: w["left"])
            rows.append(current_row)
            current_row = [word]

    # Don't forget the last row
    if current_row:
        current_row.sort(key=lambda w: w["left"])
        rows.append(current_row)

    return rows


def _split_row_into_cells(row_words: list[dict]) -> list[str]:
    """
    Split a row of words into table cells by detecting column gaps.

    Algorithm:
      1. Compute the gap between each consecutive pair of words.
      2. Determine a threshold: the larger of (median_gap × multiplier)
         and (average_word_width × factor) and a minimum pixel floor.
      3. Gaps exceeding the threshold are column separators.
      4. Words between separators are joined into a single cell.
    """
    if not row_words:
        return []

    if len(row_words) == 1:
        return [row_words[0]["text"]]

    # ── Compute gaps between consecutive words ──
    gaps: list[float] = []
    for i in range(len(row_words) - 1):
        right_edge = row_words[i]["left"] + row_words[i]["width"]
        gap = row_words[i + 1]["left"] - right_edge
        gaps.append(max(0.0, float(gap)))  # Clamp negative (overlapping words)

    if not gaps:
        # All words have zero gap — treat as single cell
        return [" ".join(w["text"] for w in row_words)]

    # ── Determine column gap threshold ──
    sorted_gaps = sorted(gaps)
    median_gap = sorted_gaps[len(sorted_gaps) // 2]

    avg_word_width = sum(w["width"] for w in row_words) / len(row_words)

    threshold = max(
        median_gap * _COLUMN_GAP_MULTIPLIER,
        avg_word_width * _COLUMN_GAP_WORD_WIDTH_FACTOR,
        _MIN_COLUMN_GAP_PX,
    )

    logger.debug(
        "Row gap analysis: %d words, median_gap=%.1f, avg_width=%.1f, threshold=%.1f",
        len(row_words), median_gap, avg_word_width, threshold,
    )

    # ── Split into cells ──
    cells: list[str] = []
    current_cell_words = [row_words[0]["text"]]

    for i, gap in enumerate(gaps):
        if gap > threshold:
            # Column boundary detected — finalize current cell
            cells.append(" ".join(current_cell_words))
            current_cell_words = [row_words[i + 1]["text"]]
        else:
            # Same cell — append word
            current_cell_words.append(row_words[i + 1]["text"])

    # Finalize last cell
    if current_cell_words:
        cells.append(" ".join(current_cell_words))

    return cells


# ═══════════════════════════════════════════════════════════════════════════════
# Excel Writing
# ═══════════════════════════════════════════════════════════════════════════════

def _write_excel(table: list[list[str]], output_path: Path) -> int:
    """
    Write a 2D table to an Excel file with formatting.

    The first row is styled as a header if it looks like one (all cells
    non-empty and generally shorter than data rows — heuristic).

    Column widths are auto-sized based on content length.

    Returns:
        Number of data rows (excluding header if present).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Extracted Data"

    # ── Write data ──
    for row_idx, row_data in enumerate(table, start=1):
        for col_idx, cell_value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # ── Detect and format header row ──
    has_header = _looks_like_header(table)
    if has_header and table:
        header_fill = PatternFill(
            start_color=_HEADER_FILL_COLOR,
            end_color=_HEADER_FILL_COLOR,
            fill_type="solid",
        )
        header_font = Font(
            bold=True,
            color=_HEADER_FONT_COLOR,
            size=11,
        )
        for col_idx in range(1, len(table[0]) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                vertical="center",
                horizontal="center",
                wrap_text=True,
            )

    # ── Auto-size column widths ──
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=1):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

        # Add padding and cap at maximum
        adjusted_width = min(max_length + 3, _MAX_COLUMN_WIDTH_CHARS)
        # Minimum width of 8 for empty columns
        adjusted_width = max(adjusted_width, 8)
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col_idx)
        ].width = adjusted_width

    # ── Add thin borders to all data cells ──
    from openpyxl.styles import Border, Side
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border

    # ── Save ──
    wb.save(str(output_path))

    # Return data row count (excluding header)
    data_rows = len(table) - 1 if has_header else len(table)
    return max(0, data_rows)


def _looks_like_header(table: list[list[str]]) -> bool:
    """
    Heuristic: the first row is a header if ALL of its cells are non-empty
    and the average cell length in the first row is shorter than the average
    of subsequent rows (headers tend to be short labels like "Name", "Date").
    """
    if len(table) < 2:
        return False

    first_row = table[0]
    remaining_rows = table[1:]

    # All first-row cells must be non-empty
    if not all(cell.strip() for cell in first_row):
        return False

    # Need at least 2 columns to look like a header
    if len(first_row) < 2:
        return False

    # Compare average cell length
    first_avg = sum(len(c) for c in first_row) / len(first_row)
    remaining_lengths = [
        len(c)
        for row in remaining_rows
        for c in row
        if c.strip()
    ]

    if not remaining_lengths:
        return False

    remaining_avg = sum(remaining_lengths) / len(remaining_lengths)

    # Header cells are typically shorter than data cells
    return first_avg < remaining_avg


# ═══════════════════════════════════════════════════════════════════════════════
# Tesseract Verification
# ═══════════════════════════════════════════════════════════════════════════════

def _verify_tesseract() -> None:
    """
    Verify that Tesseract is accessible.

    If a custom path is configured, set it before checking.
    This is called once per conversion (not once per module import)
    so the path can be changed at runtime if needed.
    """
    if TESSERACT_PATH != "tesseract":
        pytesseract.pytesseract.tesseract_cmd = TESSERACT_PATH
